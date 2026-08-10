import csv
import io
import logging
import re

import httpx
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from models import Sheet, User

logger = logging.getLogger(__name__)

REQUEST_TIMEOUT = 20
MAX_CHARS = 15000                 # flattened sheet data kept for LLM context
MAX_ROWS = 200                    # rows flattened per sheet (protects context window)
MAX_RECENT_SHEETS = 3

GOOGLE_SHEETS_EXPORT_URL = "https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
SHEET_ID_RE = re.compile(r"spreadsheets/d/([a-zA-Z0-9_-]+)")


def _flatten(rows: list[list]) -> str:
    """Convert a table into readable text lines."""
    lines = []
    for row in rows:
        cells = [
            str(c).replace("\n", " ").strip() if c is not None else ""
            for c in row
        ]
        lines.append(" | ".join(cells).rstrip(" |"))
    return "\n".join(lines)


def parse_csv_text(content: str) -> str:
    """Parse CSV text into flattened rows."""
    rows = list(csv.reader(io.StringIO(content)))
    rows = [row for row in rows if any(c.strip() for c in row)]
    truncated = len(rows) > MAX_ROWS
    text = _flatten(rows[:MAX_ROWS])
    if truncated:
        text += f"\n[... {len(rows) - MAX_ROWS} more rows truncated]"
    return text


def parse_spreadsheet(path: str, filename: str) -> str:
    """Parse a CSV or XLSX file into flattened text. Returns '' if unsupported."""
    name = filename.lower()
    if name.endswith(".csv"):
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            return parse_csv_text(f.read())

    if name.endswith((".xlsx", ".xlsm")):
        wb = load_workbook(path, read_only=True, data_only=True)
        parts = []
        try:
            for ws in wb.worksheets:
                rows = [
                    list(r)
                    for r in ws.iter_rows(values_only=True)
                    if any(c is not None for c in r)
                ][:MAX_ROWS]
                if not rows:
                    continue
                parts.append(f"[Sheet: {ws.title}]\n" + _flatten(rows))
        finally:
            wb.close()
        if not parts:
            return ""
        text = "\n\n".join(parts)
        return text[:MAX_CHARS]

    return ""


def extract_sheet_id(url_or_id: str) -> str | None:
    """Pull the sheet ID from a Google Sheets URL, or accept a bare ID."""
    match = SHEET_ID_RE.search(url_or_id)
    if match:
        return match.group(1)
    candidate = url_or_id.strip()
    if candidate and "/" not in candidate:
        return candidate
    return None


def fetch_google_sheet(url_or_id: str) -> str:
    """Fetch a publicly shared Google Sheet as CSV and flatten it.

    Only works for sheets shared with link access ('Anyone with the link').
    """
    sheet_id = extract_sheet_id(url_or_id)
    if not sheet_id:
        return f"ERROR: could not parse a Google Sheets link from '{url_or_id}'."
    try:
        with httpx.Client(timeout=REQUEST_TIMEOUT, follow_redirects=True) as client:
            response = client.get(GOOGLE_SHEETS_EXPORT_URL.format(sheet_id=sheet_id))
            response.raise_for_status()
        if "text/html" in (response.headers.get("content-type", "") or ""):
            return (
                f"ERROR: the sheet '{sheet_id}' is private or the link is invalid. "
                "The sheet must be shared with 'Anyone with the link' access for me to read it."
            )
        return parse_csv_text(response.text)
    except httpx.HTTPError as exc:
        logger.warning("Google Sheets fetch failed for %s: %s", sheet_id, exc)
        return (
            f"ERROR: could not fetch Google Sheet '{sheet_id}'. Confirm the link "
            "is publicly shared, then try again."
        )


def save_sheet(session: Session, user: User, name: str, content: str) -> Sheet:
    """Store a spreadsheet's flattened data for future conversational analysis."""
    sheet = Sheet(user_id=user.id, name=name[:255], content=content[:MAX_CHARS])
    session.add(sheet)
    session.commit()
    logger.info("Saved sheet '%s' for user %s (%d chars)", name, user.id, len(content))
    return sheet


def recent_sheets(session: Session, user: User, limit: int = MAX_RECENT_SHEETS) -> list[Sheet]:
    """The user's most recently added spreadsheets, newest first."""
    return (
        session.query(Sheet)
        .filter(Sheet.user_id == user.id)
        .order_by(Sheet.created_at.desc())
        .limit(limit)
        .all()
    )


def sheet_context_block(session: Session, user: User) -> str:
    """Compile recent spreadsheets into a compact LLM context block."""
    sheets = recent_sheets(session, user)
    if not sheets:
        return ""
    blocks = [
        f"[Spreadsheet: {sheet.name}]\n{sheet.content}" for sheet in reversed(sheets)
    ]
    return (
        "The user has the following spreadsheets in context (rows separated by '|'). "
        "Use them to answer questions about KPIs, trends, anomalies, forecasts, or "
        "comparisons. If the question is unrelated to these sheets, ignore them.\n\n"
        + "\n\n---\n\n".join(blocks)
    )